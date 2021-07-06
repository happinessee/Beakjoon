#!/usr/bin/env python
# coding: utf-8

# 초기 모듈 불러옴
import binance_d
import binance_f
import pandas as pd
import numpy as np
import os
import datetime
import decimal
import openpyxl
import requests
import time

from openpyxl import Workbook
from binance.enums import *
from binance.exceptions import BinanceAPIException, BinanceOrderException
from binance.client import Client
from binance_f import RequestClient
from binance_f.constant.test import *
from binance_f.base.printobject import *
from binance_f.constant.system import RestApiDefine
from binance_f.impl.restapirequestimpl import RestApiRequestImpl
from binance_f.impl.restapiinvoker import call_sync
from binance_f.model.constant import *


# API KEY 설정 (효혁)

binance_api= 'vDiLBxzA2VPlSuBLYXiDF47mScyldTANFIdysXlZYbVtcM9MA7g65F0EKPwG1w9k'
binance_secret= '3gs0sLJ0HnLsHBmcnqEKH46nEklakikZubDefJEbK8wOusHA8XiDtYj0Xw6LrnLY'

# slack api key (경민이꺼)
myToken = "xoxb-2216792091655-2231529462418-t7qBRPN5xN0W42Wsm0LqhBhn"


request_client = RequestClient(api_key = binance_api, secret_key = binance_secret)
client = Client(binance_api, binance_secret)

mode_Choice = request_client.get_position_v2();
account = request_client.get_account_information()


# 개인이 선택하는 변수들

# 코인 집합
symbol_set = ['ADAUSDT', 'BNBUSDT', 'ETCUSDT', 'DOGEUSDT', 
          'TRXUSDT', 'LINKUSDT', 'LTCUSDT', 'EOSUSDT', 'BCHUSDT', 'DOTUSDT']

# 코인에 따른 최대 레버리지
leverage_set = [75, 75, 75, 50, 75, 75, 75, 75, 75, 75]

# 코인에 따라 허락되는 정밀도 (소수점 자리)
amount = [0, 2, 2, 0, 0, 2, 3, 1, 2, 1]

# 우리가 넣은 초기 USDT 양
initial_entry_usdt = 0.20

# usdt 조정을 위한 count (usdt양이 100 추가 될 때마다 1씩 증가)
count = 1

# 펀딩비 관련-----------------------------------------------------------------------

# 초기변수 c 값
c = -300

# 물타기, 불타기 횟수
n = 3

# 수식 계산 전 기본 물타기 퍼센티지
a = -50


# 코인의 가격정보를 빼와 우리가 구매할 코인의 양을 구한다.
# 매개변수로 코인의 집합, 레버리지 집합, 초기 투입 usdt량을 넣어주면 각 코인당 넣어야 할 코인의 양을 구해준다.
# 사용법
# need_quantity = getCoinQuantity(symbol_set = symbol_set, leverage_set = leverage_set, initial_entry_usdt = initial_entry_usdt) 

def get_CoinQuantity(symbol_set, leverage_set, initial_entry_usdt) :
    price_temp = []
    need_quantity = []
    coin_price_list=[]
    # 가격 정보를 가져와 가격을 뽑아내고, 필요한 코인 양을 계산해준다.
    for i in range(len(symbol_set)) :
        
        price_temp.append(request_client.get_symbol_price_ticker(symbol=symbol_set[i]))

        coin_price_list.append(price_temp[i][0].price)

        need_quantity.append(((leverage_set[i] * initial_entry_usdt) / coin_price_list[i]))
    
    # 코인의 API계산 정밀도
    amount = [0, 2, 2, 0, 0, 2, 3, 1, 2, 1]
    
    # 코인의 정밀도를 반올림시켜 맞추어준다.
    for i in range(len(symbol_set)) :
        if (amount[i] == 0) :
            need_quantity[i] = round(need_quantity[i])
        else :
            need_quantity[i] = round(need_quantity[i], amount[i])
            
    return need_quantity

# slack message 보내기
def post_message(token, channel, text):
    response = requests.post("https://slack.com/api/chat.postMessage",
        headers={"Authorization": "Bearer "+token},
        data={"channel": channel,"text": text}
    )
    print(response)

# 잔고에 따른 initial_entry_usdt 조정 (0.5%)
def modify_usdt (totalMarginBalance, initial_entry_usdt, count) :
    
    temp1 = 100
    temp2 = 0.5

    if (totalMarginBalance >= temp1 * count) :
        temp2 = temp2 * count
        count += 1
        return temp2
    
    else :
        return initial_entry_usdt


# 펀딩비에 따른 물타기, 불타기 조정
# long과 short의 계산식이 다르다.

# 물타기, 불타기 ROE 값을 구한다.
def get_longValue (symbol_set, a, n) :
    
    # coin 당 펀딩비를 넣어줄 임시 리스트
    funding_rate = []
    
    # 단순 count
    count0 = 0

    # 코인에 따라 달라지는 b값 리스트
    b = []
    
    value = []

    for i in range(len(symbol_set)) :

        funding_rate.append(request_client.get_funding_rate(symbol=symbol_set[i], limit = 1))
        b.append(((funding_rate[count0][0].fundingRate - 0.0001) * c) + 1)    # 해당 공식을 적용시켜준다.
        count0 += 1
        
    for i in range(len(symbol_set)) :
        temp = []

        for j in range (n + 1) :
            temp.append(a * b[i] * (2 ** (j)))
        value.append(temp)
    return value


def get_shortValue (symbol_set, a, n) :
    
    # coin 당 펀딩비를 넣어줄 임시 리스트
    funding_rate = []
    
    # 단순 count
    count0 = 0

    # 코인에 따라 달라지는 b값 리스트
    b = []
    
    value = []

    for i in range(len(symbol_set)) :

        funding_rate.append(request_client.get_funding_rate(symbol=symbol_set[i], limit = 1))
        b.append(((funding_rate[count0][0].fundingRate - 0.0001) * c) + 1)    # 해당 공식을 적용시켜준다.
        count0 += 1
    
    for i in range(len(symbol_set)) :
        temp = []
        for j in range (n + 1) :
            temp.append(a * b[i] * 2 * (2 ** (j)))
        value.append(temp)

        
    return value

# 현재 ROE 값을 계산
def get_CurrentROE(symbol_set, initial_entry_usdt, total_pnl) :
    
    ROE = []
    
    for i in range(len(symbol_set)) :
        ROE.insert(i, (total_pnl[i] / initial_entry_usdt) * 100)
    
    return ROE

# 실제 동작 알고리즘에서 판단할 각 포지션 당 ROE
def get_positionROE(symbol_set, initial_entry_usdt, mode_Choice) :
    
    ROE = []
    total_pnl_temp = []
    for i in (mode_Choice) :
        for j in symbol_set :
            if (i.symbol == j) :
                total_pnl_temp.append(i.unrealizedProfit)
    
    for i in range(len(total_pnl_temp)) :
        ROE.append((total_pnl_temp[i] / initial_entry_usdt) * 100)
        
    return ROE


# totalpnl을 계산해 우리는 totalpnl이 initial_entry_usdt를 넘으면 포지션을 종료하고 다시 재진입하기로 했었다.
# 그것을 위한 TOTALPNL 계산이다.
def get_total_pnl(symbol_set, mode_Choice) :
    
    total_pnl_temp = []
    # pnl을 배열로 가져온다.
    
    for i in (mode_Choice) :
        for j in symbol_set :
            
            if (i.symbol == j) :
                total_pnl_temp.append(i.unrealizedProfit)

    total_pnl = []

    # 양 포지션 (long, short)의 pnl을 더해 total_pnl로 만듬
    for i in range (len(total_pnl_temp)) : # length 18
        if (i % 2 == 0) :
            total_pnl.append(total_pnl_temp[i] + total_pnl_temp[i+1])
            
    return total_pnl

# 넣을 코인의 정보 가져오기---------------------------------------
def get_Coinprice(symbol_set) :
    price_temp = []
    for i in symbol_set :
        price_temp.append(request_client.get_symbol_price_ticker(symbol=i))


    # 코인의 정보에서 가격 정보만 빼서 저장
    coin_price = {}        # coin_price['BTCUSDT'] 형식으로 가격을 꺼내쓸 수 있게 만들어뒀다.
    coin_price_list = []

    count1 = 0     # 밑의 반복문 횟수 count
    for i in range (len(symbol_set)) :
        coin_price[symbol_set[i]] = price_temp[i][0].price
        coin_price_list.append(price_temp[i][0].price)
        count1 += 1
        
    return coin_price, coin_price_list


# 엑셀 파일로 로그 만들기.

def save_log (mode_Choice, side, positionSide, need_quantity, count_watering, count_firing, symbol, coin_log_dir) :
    
    # 거래한 코인 총 양 표시
    total_amount = []
    
    # 시간 표시
    now = datetime.datetime.now()
    nowDatetime = now.strftime('%Y-%m-%d %H:%M:%S')
    
    # 코인 포지션당 총 양을 구해줌
    for j in mode_Choice :
        if(j.symbol == symbol) :
            total_amount.append(j.positionAmt)
    
    # 만든 엑셀 파일을 연다.
    write_wb = openpyxl.load_workbook(coin_log_dir)
    
    # 해당 코인의 sheet을 선택한다.
    write_ws = write_wb[symbol]
    
    # 구매, 판매 정보를 행에 넣고 저장한다.
    write_ws.append([nowDatetime, side, positionSide, need_quantity, (total_amount[0] if positionSide == 'LONG' else total_amount[1]), count_watering, count_firing])
    write_wb.save(coin_log_dir)

# 딱 한 번만 실행되어야 할 코드 ---------------------------------------------------------------

# 엑셀에 필요했던 기본 코드들. 다시 실행 x

# wb = openpyxl.load_workbook(coin_log_dir)

# for i in symbol_set :
#     ws = wb[i]
#     ws.append(['거래 시간', '구매-판매', '포지션', '구매한 코인 양', '포지션에 보유중인 코인 양', '물타기 횟수', '불타기 횟수'])
    
# wb.save(coin_log_dir)

post_message(myToken, "#projec", "프로그램 작동이 시작되었습니다!")
need_quantity = get_CoinQuantity(symbol_set = symbol_set, leverage_set = leverage_set, initial_entry_usdt = initial_entry_usdt)
total_amount = []

part = "0-0"

# coin_log_dir = '/home/ubuntu/Project/Binance_Futures_python-master/Coin_Log.xlsx'

for i in range(len(symbol_set)) :
     client.futures_change_leverage(symbol = symbol_set[i], leverage = leverage_set[i])
    
     client.futures_create_order(symbol=symbol_set[i], side='BUY', positionSide = 'LONG', type='MARKET', quantity=need_quantity[i])
     client.futures_create_order(symbol=symbol_set[i], side='SELL', positionSide = 'SHORT', type='MARKET', quantity=need_quantity[i])

part =  "0-1"

# 코인 별 물, 불탄 횟수
count_watering = []
count_firing = []

# 코인마다 물타기, 불타기 횟수를 기록하기 위해 0으로 초기화.
for i in range (len(symbol_set)) :
    count_watering.append(0)
    count_firing.append(0)
    
for j in mode_Choice :
    for k in symbol_set :
        if (j.symbol == k) :
            total_amount.append(j.positionAmt)

for i in range(len(symbol_set)) :
    print(total_amount[i])
# 데이터 딕셔너리를 만들어 엑셀에 저장해 주기 위한 코드
# for i in range(len(symbol_set)) :
    
#     save_log(mode_Choice = mode_Choice, side = 'OPEN', positionSide = 'LONG', need_quantity = need_quantity[i],
#                           count_watering = count_watering[i], count_firing = count_firing[i], symbol = symbol_set[i], coin_log_dir = coin_log_dir)
#     save_log(mode_Choice = mode_Choice, side = 'OPEN', positionSide = 'SHORT', need_quantity = need_quantity[i],
#                           count_watering = count_watering[i], count_firing = count_firing[i], symbol = symbol_set[i], coin_log_dir = coin_log_dir)
part = "0-2"

# 계속 반복해서 실행시켜 주어야 하는, 실시간으로 초기화 해야 하는 코드 

try :
    while(True) :
        
    # 계속 초기화 해야 하는 변수 --------------------------------------------
        part = "1-0"

        # 클라이언트의 종합 정보를 가져오는 변수
        mode_Choice = request_client.get_position_v2()
        # 계좌의 정보를 가져오는 변수
        account = request_client.get_account_information()

        part = "1-1"

    # -----------------------------------------------------------------------
        # 구매할 코인의 양 초기화
        need_quantity = get_CoinQuantity(symbol_set = symbol_set, leverage_set = leverage_set, initial_entry_usdt = initial_entry_usdt)

        part = "1-2"

        # total_pnl을 구해준다. (초기 pnl 합은 약 0이다.)
        total_pnl = get_total_pnl(symbol_set = symbol_set, mode_Choice = mode_Choice)

        part = "1-3"

        # 현재 ROE를 구해준다.
        current_ROE = get_positionROE(symbol_set = symbol_set, initial_entry_usdt = initial_entry_usdt, mode_Choice = mode_Choice)

        part = "1-4"

        # long일 때, 거래되는 ROE값, short일 때 거래되는 ROE 값
        long_roe_value = get_longValue(symbol_set = symbol_set, a = a, n = n)
        short_roe_value = get_shortValue(symbol_set = symbol_set, a = a, n= n)

        part = "1-5"

        # usdt 양을 100달러가 추가될 때마다 재조정
        initial_entry_usdt = modify_usdt(totalMarginBalance = account.totalMarginBalance, initial_entry_usdt = initial_entry_usdt, count = count)

        for i in range (len(symbol_set)) :

            part = "2-1"

            # 롱 포지션이 손해 (long 포지션이 (-)%)
            if (current_ROE[2 * i] <= long_roe_value[i][count_watering[i]]) :

                part = "2-2"

                if (count_watering[i] >= 1 and count_firing[i] < n) :

                    client.futures_create_order(symbol=symbol_set[i], side='SELL', positionSide = 'SHORT', type='MARKET', quantity=need_quantity[i])
                    count_firing[i] += 1
                    
                    # save_log (mode_Choice = mode_Choice, side = 'OPEN', positionSide = 'SHORT', need_quantity = need_quantity[i],
                    #         count_watering = count_watering[i], count_firing = count_firing[i], symbol = symbol_set[i], coin_log_dir = coin_log_dir)

                part = "2-3"

                if (count_watering[i] < n) :

                    client.futures_create_order(symbol=symbol_set[i], side='BUY', positionSide = 'LONG', type='MARKET', quantity=need_quantity[i])
                    count_watering[i] += 1
                    
                    # save_log(mode_Choice = mode_Choice, side = 'OPEN', positionSide = 'LONG', need_quantity = need_quantity[i],
                    #         count_watering = count_watering[i], count_firing = count_firing[i], symbol = symbol_set[i], coin_log_dir = coin_log_dir)
            
                part = "3-1"

            # 숏 포지션이 손해 
            elif (current_ROE[2 * i + 1] <= short_roe_value[i][count_watering[i]]) :
                part = "3-2"

                if (count_watering[i] >= 1 and count_firing[i] < n) :

                    client.futures_create_order(symbol=symbol_set[i], side='BUY', positionSide = 'LONG', type='MARKET', quantity=need_quantity[i])
                    count_firing[i] += 1
                    
                    # save_log(mode_Choice = mode_Choice, side = 'OPEN', positionSide = 'LONG', need_quantity = need_quantity[i],
                    #         count_watering = count_watering[i], count_firing = count_firing[i], symbol = symbol_set[i], coin_log_dir = coin_log_dir)
                
                part = "3-3"
                if (count_watering[i] < n) :

                    client.futures_create_order(symbol=symbol_set[i], side='SELL', positionSide = 'SHORT', type='MARKET', quantity=need_quantity[i])
                    count_watering[i] += 1
                    
                    # save_log (mode_Choice = mode_Choice, side = 'OPEN', positionSide = 'SHORT', need_quantity = need_quantity[i],
                    #         count_watering = count_watering[i], count_firing = count_firing[i], symbol = symbol_set[i], coin_log_dir = coin_log_dir)

            # 총 pnl이 init ial_entry_usdt(0.15)보다 커진다면 모든 포지션 종료 후 다시 재진입
            if (total_pnl[i] > initial_entry_usdt) :
                part = "4-1"
                
                for j in mode_Choice :
                    for k in symbol_set :
                        if (j.symbol == k) :
                            total_amount.append(j.positionAmt)         
                        
                client.futures_create_order(symbol=symbol_set[i], side='SELL', positionSide = 'LONG', type='MARKET', quantity=total_amount[2*i])
                
                # save_log(mode_Choice = mode_Choice, side = 'CLOSE', positionSide = 'LONG', need_quantity = total_amount[2*i],
                #             count_watering = 'COIN_SELL', count_firing = 'COIN_SELL', symbol = symbol_set[i], coin_log_dir = coin_log_dir)
                
                client.futures_create_order(symbol=symbol_set[i], side='BUY', positionSide = 'SHORT', type='MARKET', quantity=abs(total_amount[2*i + 1]))
                
                # save_log(mode_Choice = mode_Choice, side = 'CLOSE', positionSide = 'SHORT', need_quantity = total_amount[2*i + 1],
                #             count_watering = 'COIN_SELL', count_firing = 'COIN_SELL', symbol = symbol_set[i], coin_log_dir = coin_log_dir)
                
                account = request_client.get_account_information()
                message = (symbol_set[i] + "를 거래했습니다.") + ("\n미실현 손익 :  %f" % account.totalUnrealizedProfit) + ("\n현재 지갑 잔고 : %f" % account.totalWalletBalance)
                post_message(myToken, "#projec", message)

                part = "4-2"

                # 재진입
                client.futures_create_order(symbol=symbol_set[i], side='BUY', positionSide = 'LONG', type='MARKET', quantity=need_quantity[i])
                client.futures_create_order(symbol=symbol_set[i], side='SELL', positionSide = 'SHORT', type='MARKET', quantity=need_quantity[i])
                
                part = "4-3"
                count_watering[i] = 0
                count_firing[i] = 0
                
                # save_log(mode_Choice = mode_Choice, side = 'OPEN', positionSide = 'LONG', need_quantity = need_quantity[i],
                #             count_watering = count_watering[i], count_firing = count_firing[i], symbol = symbol_set[i], coin_log_dir = coin_log_dir)
                
                # save_log(mode_Choice = mode_Choice, side = 'OPEN', positionSide = 'SHORT', need_quantity = need_quantity[i],
                #             count_watering = count_watering[i], count_firing = count_firing[i], symbol = symbol_set[i], coin_log_dir = coin_log_dir)

except :
    error_message = part + "부분에서 에러가 발생했습니다."
    post_message(myToken, "#projec", error_message)
    post_message(myToken, "#projec", "프로그램이 오류로 인해 종료되었습니다.. 확인해주세요")
       
# -------------------------------------------------------------------------------------------------------------

# coin_log_dir = 'C:\\Users\\gygur\\Desktop\\Coin_Log.xlsx'

# # 엑셀파일을 '생성'해 코인 이름을 파일로 저장 = 이미 완료 했기 때문에 주석처리. (다시 실행하면 로그 파일이 사라지고 새로 덮여 씌워져버림)

# write_wb = Workbook()
# write_ws = write_wb.active

# for i in (symbol_set) :

#     createSheet = write_wb.create_sheet(i)
    
# write_wb.save(coin_log_dir)

# ----------------------------------------------------------------------------------